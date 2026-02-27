"""
ELOMER — Efficient LOgistics of a MEdical Representative
Полностью переработанный UI — стиль аналитического дашборда.
"""

import sys, os, pandas as pd
from PySide6.QtWidgets import *
from PySide6.QtCore    import *
from PySide6.QtGui     import *
from PySide6.QtWebEngineWidgets import QWebEngineView
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
import matplotlib.pyplot as plt
import io, numpy as np, random, math, time
from datetime import datetime, timedelta

from calculator_core import MedicalRepCalculatorGUI

# ─────────────────────────────────────────────────────────────────────────────
# ЦВЕТОВЫЕ ТЕМЫ
# ─────────────────────────────────────────────────────────────────────────────

DARK_THEME = {
    'bg':          '#0D1117',
    'bg2':         '#161B27',
    'surface':     '#1C2333',
    'surface2':    '#232B3E',
    'surface3':    '#2D3855',
    'panel':       '#161B27',
    'accent':      '#5B8FF9',
    'accent2':     '#3A6FE4',
    'success':     '#36CFA0',
    'success2':    '#10B981',
    'warning':     '#F7B731',
    'warning2':    '#E0A820',
    'danger':      '#FC5C65',
    'danger2':     '#E03E47',
    'purple':      '#A77DFA',
    'purple2':     '#8B5CF6',
    'cyan':        '#26D4EF',
    'text':        '#E2E8F4',
    'text2':       '#8B98B8',
    'text3':       '#445070',
    'border':      '#1E293D',
    'border2':     '#2A3A58',
    'plt_bg':      '#1C2333',
    'plt_surf':    '#232B3E',
    'plt_text':    '#8B98B8',
    'plt_grid':    '#1E293D',
    'plt_blue':    '#5B8FF9',
    'plt_green':   '#36CFA0',
    'plt_orange':  '#F7B731',
    'plt_red':     '#FC5C65',
    'plt_purple':  '#A77DFA',
}

LIGHT_THEME = {
    'bg':          '#F0F4FA',
    'bg2':         '#E4EAF4',
    'surface':     '#FFFFFF',
    'surface2':    '#F5F8FF',
    'surface3':    '#EBF0FA',
    'panel':       '#FFFFFF',
    'accent':      '#2563EB',
    'accent2':     '#1D4ED8',
    'success':     '#059669',
    'success2':    '#047857',
    'warning':     '#D97706',
    'warning2':    '#B45309',
    'danger':      '#DC2626',
    'danger2':     '#B91C1C',
    'purple':      '#7C3AED',
    'purple2':     '#6D28D9',
    'cyan':        '#0891B2',
    'text':        '#0F172A',
    'text2':       '#475569',
    'text3':       '#94A3B8',
    'border':      '#E2E8F0',
    'border2':     '#CBD5E1',
    'plt_bg':      '#FFFFFF',
    'plt_surf':    '#F5F8FF',
    'plt_text':    '#475569',
    'plt_grid':    '#E2E8F0',
    'plt_blue':    '#2563EB',
    'plt_green':   '#059669',
    'plt_orange':  '#D97706',
    'plt_red':     '#DC2626',
    'plt_purple':  '#7C3AED',
}

C = dict(DARK_THEME)
_current_theme = 'dark'

def set_theme(name: str):
    global _current_theme
    _current_theme = name
    C.clear()
    C.update(DARK_THEME if name == 'dark' else LIGHT_THEME)

def _stylesheet() -> str:
    return f"""
QMainWindow {{ background-color: {C['bg']}; }}
QWidget {{ background-color: transparent; color: {C['text']};
           font-family: "Segoe UI", "Inter", Arial, sans-serif; font-size: 13px; }}
QScrollArea {{ border: none; background-color: transparent; }}
QScrollBar:vertical {{ background: transparent; width: 5px; margin: 0; border-radius: 3px; }}
QScrollBar::handle:vertical {{ background: {C['border2']}; min-height: 30px; border-radius: 3px; }}
QScrollBar::handle:vertical:hover {{ background: {C['accent']}; }}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{ height: 0; }}
QScrollBar:horizontal {{ background: transparent; height: 5px; border-radius: 3px; }}
QScrollBar::handle:horizontal {{ background: {C['border2']}; min-width: 30px; border-radius: 3px; }}
QScrollBar::handle:horizontal:hover {{ background: {C['accent']}; }}
QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {{ width: 0; }}
QStatusBar {{ background-color: {C['bg2']}; color: {C['text3']};
              border-top: 1px solid {C['border']}; font-size: 11px; padding: 0 16px; min-height: 28px; }}
QMenuBar {{ background-color: {C['bg2']}; color: {C['text2']};
            border-bottom: 1px solid {C['border']}; padding: 0 8px; min-height: 32px; }}
QMenuBar::item {{ padding: 6px 12px; border-radius: 5px; margin: 2px; }}
QMenuBar::item:selected {{ background-color: {C['surface2']}; color: {C['text']}; }}
QMenu {{ background-color: {C['surface']}; border: 1px solid {C['border2']};
         border-radius: 10px; padding: 5px; }}
QMenu::item {{ padding: 8px 20px; border-radius: 6px; color: {C['text2']}; }}
QMenu::item:selected {{ background-color: {C['accent']}; color: white; }}
QMenu::separator {{ height: 1px; background: {C['border']}; margin: 3px 8px; }}
QToolTip {{ background-color: {C['surface2']}; color: {C['text']};
            border: 1px solid {C['border2']}; border-radius: 8px; padding: 6px 10px; }}
QDialog, QMessageBox {{ background-color: {C['surface']}; }}
"""

# ─────────────────────────────────────────────────────────────────────────────
# ВИДЖЕТЫ
# ─────────────────────────────────────────────────────────────────────────────

class AppButton(QPushButton):
    def __init__(self, text, parent=None, variant='primary'):
        super().__init__(text, parent)
        self.variant = variant
        self.setSizePolicy(QSizePolicy.MinimumExpanding, QSizePolicy.Fixed)
        self.setFixedHeight(40)
        self.setCursor(Qt.PointingHandCursor)
        self._apply_style()

    def _apply_style(self):
        v = self.variant
        styles = {
            'primary':   (C['accent'],   C['accent2'],   '#FFFFFF', 'none'),
            'success':   (C['success'],  C['success2'],  '#FFFFFF', 'none'),
            'danger':    (C['danger'],   C['danger2'],   '#FFFFFF', 'none'),
            'warning':   (C['warning'],  C['warning2'],  '#FFFFFF', 'none'),
            'ghost':     ('transparent', C['surface2'],  C['text2'], f'1px solid {C["border2"]}'),
            'secondary': (C['surface2'], C['surface3'],  C['text'],  f'1px solid {C["border2"]}'),
        }
        bg, hover, tc, border = styles.get(v, styles['primary'])
        self.setStyleSheet(f"""
            QPushButton {{
                background-color: {bg}; color: {tc};
                border: {border}; padding: 0 18px; border-radius: 10px;
                font-size: 13px; font-weight: 600;
            }}
            QPushButton:hover {{ background-color: {hover}; }}
            QPushButton:pressed {{ padding: 1px 18px 0 18px; }}
            QPushButton:disabled {{ background-color: {C['surface2']}; color: {C['text3']}; border: 1px solid {C['border']}; }}
        """)

    def refresh_theme(self):
        self._apply_style()


class AppComboBox(QComboBox):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFixedHeight(40)
        self.refresh_theme()

    def refresh_theme(self):
        self.setStyleSheet(f"""
            QComboBox {{
                padding: 0 36px 0 13px; border: 1px solid {C['border2']};
                border-radius: 10px; background-color: {C['surface2']};
                color: {C['text']}; font-size: 13px;
            }}
            QComboBox:hover {{ border-color: {C['accent']}; background-color: {C['surface3']}; }}
            QComboBox:focus {{ border: 2px solid {C['accent']}; }}
            QComboBox::drop-down {{ border: none; width: 32px; }}
            QComboBox::down-arrow {{
                image: none; border-left: 4px solid transparent;
                border-right: 4px solid transparent; border-top: 5px solid {C['text2']};
                margin-right: 10px;
            }}
            QComboBox QAbstractItemView {{
                background-color: {C['surface']}; border: 1px solid {C['border2']};
                border-radius: 10px; selection-background-color: {C['accent']};
                color: {C['text']}; padding: 4px; outline: none;
            }}
        """)


class AppSpinBox(QSpinBox):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFixedHeight(40)
        self.refresh_theme()

    def refresh_theme(self):
        self.setStyleSheet(f"""
            QSpinBox {{
                padding: 0 8px; border: 1px solid {C['border2']};
                border-radius: 10px; background-color: {C['surface2']};
                color: {C['text']}; font-size: 13px;
            }}
            QSpinBox:hover {{ border-color: {C['accent']}; background-color: {C['surface3']}; }}
            QSpinBox:focus {{ border: 2px solid {C['accent']}; }}
            QSpinBox::up-button, QSpinBox::down-button {{
                border: none; background: transparent; width: 22px;
            }}
            QSpinBox::up-button:hover, QSpinBox::down-button:hover {{ background: {C['surface3']}; }}
            QSpinBox::up-arrow {{ border-left: 4px solid transparent; border-right: 4px solid transparent; border-bottom: 5px solid {C['text2']}; }}
            QSpinBox::down-arrow {{ border-left: 4px solid transparent; border-right: 4px solid transparent; border-top: 5px solid {C['text2']}; }}
        """)


class AppCheckBox(QCheckBox):
    def __init__(self, text, parent=None):
        super().__init__(text, parent)
        self.refresh_theme()

    def refresh_theme(self):
        self.setStyleSheet(f"""
            QCheckBox {{ spacing: 10px; font-size: 13px; color: {C['text2']}; padding: 4px 0; }}
            QCheckBox::indicator {{
                width: 18px; height: 18px; border: 2px solid {C['border2']};
                border-radius: 5px; background-color: {C['surface2']};
            }}
            QCheckBox::indicator:checked {{ background-color: {C['accent']}; border-color: {C['accent']}; }}
            QCheckBox::indicator:hover {{ border-color: {C['accent']}; }}
        """)


class FormField(QWidget):
    """Label сверху + widget снизу — без наложений"""
    def __init__(self, label_text, widget, parent=None):
        super().__init__(parent)
        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(5)
        self._lbl = QLabel(label_text)
        self._style_label()
        lay.addWidget(self._lbl)
        lay.addWidget(widget)

    def _style_label(self):
        self._lbl.setStyleSheet(
            f"color: {C['text3']}; font-size: 10px; font-weight: 700; letter-spacing: 1px;")

    def refresh_theme(self):
        self._style_label()


class SectionDivider(QWidget):
    def __init__(self, text, parent=None):
        super().__init__(parent)
        self.setFixedHeight(30)
        lay = QHBoxLayout(self)
        lay.setContentsMargins(0, 10, 0, 0)
        lay.setSpacing(8)
        self._lbl = QLabel(text.upper())
        self._style_label()
        lay.addWidget(self._lbl)
        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setFixedHeight(1)
        self._line = line
        self._style_line()
        lay.addWidget(line, stretch=1)

    def _style_label(self):
        self._lbl.setStyleSheet(
            f"color: {C['text3']}; font-size: 10px; font-weight: 700; letter-spacing: 1.5px;")

    def _style_line(self):
        self._line.setStyleSheet(f"background-color: {C['border']}; color: {C['border']};")

    def refresh_theme(self):
        self._style_label()
        self._style_line()


class KPICard(QFrame):
    def __init__(self, title, value, unit='', color=None, icon='●', parent=None):
        super().__init__(parent)
        self._color = color or C['accent']
        self.setMinimumWidth(120)
        self.setStyleSheet(f"""
            QFrame {{
                background-color: {C['surface']};
                border-radius: 14px;
                border: 1px solid {C['border']};
            }}
        """)
        lay = QVBoxLayout(self)
        lay.setContentsMargins(14, 12, 14, 12)
        lay.setSpacing(6)

        top_row = QHBoxLayout()
        top_row.setSpacing(7)

        icon_lbl = QLabel(icon)
        icon_lbl.setFixedSize(28, 28)
        icon_lbl.setAlignment(Qt.AlignCenter)
        icon_lbl.setStyleSheet(f"""
            background-color: {self._color}22; color: {self._color};
            border-radius: 7px; font-size: 13px;
        """)
        top_row.addWidget(icon_lbl)

        t_lbl = QLabel(title)
        t_lbl.setStyleSheet(f"color: {C['text2']}; font-size: 11px;")
        top_row.addWidget(t_lbl)
        top_row.addStretch()
        lay.addLayout(top_row)

        val_row = QHBoxLayout()
        val_row.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        val_row.setSpacing(4)
        vl = QLabel(str(value))
        vl.setStyleSheet(
            f"color: {self._color}; font-size: 22px; font-weight: 700; letter-spacing: -0.5px;")
        val_row.addWidget(vl)
        if unit:
            ul = QLabel(unit)
            ul.setStyleSheet(f"color: {C['text2']}; font-size: 12px; padding-top: 6px;")
            val_row.addWidget(ul)
        lay.addLayout(val_row)


# Алиас для совместимости
MetricCard = KPICard


class BottomTabBar(QWidget):
    """Таббар внизу — как в референсных примерах"""
    tabChanged = Signal(int)

    def __init__(self, tabs, parent=None):
        super().__init__(parent)
        self.setFixedHeight(50)
        self._current = 0
        self._buttons = []
        lay = QHBoxLayout(self)
        lay.setContentsMargins(16, 6, 16, 6)
        lay.setSpacing(4)
        for i, (icon, label) in enumerate(tabs):
            btn = QPushButton(f"{icon}  {label}")
            btn.setFixedHeight(36)
            btn.setCursor(Qt.PointingHandCursor)
            btn.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
            btn.clicked.connect(lambda _, idx=i: self._on_tab(idx))
            self._buttons.append(btn)
            lay.addWidget(btn)
        lay.addStretch()
        self.refresh_theme()

    def _on_tab(self, idx):
        self._current = idx
        self.refresh_theme()
        self.tabChanged.emit(idx)

    def set_current(self, idx):
        self._current = idx
        self.refresh_theme()

    def refresh_theme(self):
        self.setStyleSheet(
            f"background-color: {C['bg2']}; border-top: 1px solid {C['border']};")
        for i, btn in enumerate(self._buttons):
            if i == self._current:
                btn.setStyleSheet(f"""
                    QPushButton {{
                        background-color: {C['accent']}; color: white;
                        border: none; border-radius: 9px;
                        font-size: 12px; font-weight: 600; padding: 0 14px;
                    }}
                """)
            else:
                btn.setStyleSheet(f"""
                    QPushButton {{
                        background-color: transparent; color: {C['text2']};
                        border: none; border-radius: 9px;
                        font-size: 12px; padding: 0 14px;
                    }}
                    QPushButton:hover {{ background-color: {C['surface2']}; color: {C['text']}; }}
                """)


class AppTable(QTableWidget):
    def __init__(self, columns, headers, parent=None):
        super().__init__(parent)
        self.setColumnCount(columns)
        self.setHorizontalHeaderLabels(headers)
        self.refresh_theme()
        self.setAlternatingRowColors(True)
        self.horizontalHeader().setStretchLastSection(True)
        self.verticalHeader().setVisible(False)
        self.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.setShowGrid(False)

    def refresh_theme(self):
        self.setStyleSheet(f"""
            QTableWidget {{
                background-color: {C['surface']}; alternate-background-color: {C['surface2']};
                gridline-color: {C['border']}; border: 1px solid {C['border']};
                border-radius: 12px; color: {C['text']}; font-size: 12px;
            }}
            QTableWidget::item {{ padding: 10px 12px; border: none; }}
            QTableWidget::item:selected {{ background-color: {C['accent']}33; color: {C['text']}; }}
            QHeaderView::section {{
                background-color: {C['surface2']}; color: {C['text3']};
                padding: 10px 12px; border: none;
                border-bottom: 2px solid {C['border2']};
                font-size: 10px; font-weight: 700; letter-spacing: 1px;
            }}
        """)

    def set_cell(self, row, col, text, bg_color=None, fg_color=None, bold=False, align=Qt.AlignLeft):
        item = QTableWidgetItem(str(text))
        item.setTextAlignment(align | Qt.AlignVCenter)
        if bg_color: item.setBackground(QColor(bg_color))
        if fg_color: item.setForeground(QColor(fg_color))
        if bold:
            f = item.font(); f.setBold(True); item.setFont(f)
        self.setItem(row, col, item)


# ─────────────────────────────────────────────────────────────────────────────
# ПАНЕЛИ ВВОДА
# ─────────────────────────────────────────────────────────────────────────────

class DailyCalcPanel(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._build()

    def set_calculator(self, calc):
        self.calculator = calc

    def _build(self):
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        scroll.setStyleSheet("border: none; background: transparent;")

        inner = QWidget()
        inner.setStyleSheet("background: transparent;")
        lay = QVBoxLayout(inner)
        lay.setContentsMargins(18, 18, 18, 18)
        lay.setSpacing(12)

        lay.addWidget(SectionDivider("Параметры"))

        self.city_combo = AppComboBox()
        self.city_combo.addItems(["Москва", "Санкт-Петербург", "Екатеринбург", "Новосибирск", "Казань"])
        lay.addWidget(FormField("Город", self.city_combo))

        self.spec_combo = AppComboBox()
        self.spec_combo.addItems(["Кардиологи", "Терапевты", "Педиатры", "Аптеки"])
        lay.addWidget(FormField("Специализация", self.spec_combo))

        self.transport_combo = AppComboBox()
        self.transport_combo.addItems(["Автомобиль", "Общественный транспорт", "Пешком"])
        lay.addWidget(FormField("Транспорт", self.transport_combo))

        lay.addWidget(SectionDivider("Визиты"))

        self.visits_spin = AppSpinBox()
        self.visits_spin.setRange(1, 20)
        self.visits_spin.setValue(8)
        lay.addWidget(FormField("Количество визитов", self.visits_spin))

        lay.addWidget(SectionDivider("Опции"))

        opts_card = QFrame()
        opts_card.setStyleSheet(
            f"background: {C['surface']}; border-radius: 12px; border: 1px solid {C['border']};")
        opts_lay = QVBoxLayout(opts_card)
        opts_lay.setContentsMargins(14, 12, 14, 12)
        opts_lay.setSpacing(8)
        self.show_map_checkbox = AppCheckBox("Показывать карту маршрута")
        self.use_ml_checkbox   = AppCheckBox("Использовать ML-оптимизацию")
        opts_lay.addWidget(self.show_map_checkbox)
        opts_lay.addWidget(self.use_ml_checkbox)
        lay.addWidget(opts_card)

        lay.addSpacing(6)

        self.calculate_btn = AppButton("▶   Рассчитать рабочий день", variant='primary')
        lay.addWidget(self.calculate_btn)

        self.export_btn = AppButton("⬇   Экспорт в Excel", variant='ghost')
        self.export_btn.setEnabled(False)
        lay.addWidget(self.export_btn)

        lay.addStretch()
        scroll.setWidget(inner)
        outer.addWidget(scroll)

    def refresh_theme(self):
        self.setStyleSheet(f"background-color: {C['panel']};")
        for w in self.findChildren(AppComboBox):    w.refresh_theme()
        for w in self.findChildren(AppSpinBox):     w.refresh_theme()
        for w in self.findChildren(AppCheckBox):    w.refresh_theme()
        for w in self.findChildren(AppButton):      w.refresh_theme()
        for w in self.findChildren(FormField):      w.refresh_theme()
        for w in self.findChildren(SectionDivider): w.refresh_theme()
        for f in self.findChildren(QFrame):
            if hasattr(f, 'layout') and f.layout() and f.layout().count() > 0:
                f.setStyleSheet(
                    f"background: {C['surface']}; border-radius: 12px; border: 1px solid {C['border']};")


class ProjectCalcPanel(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._build()

    def set_calculator(self, calc):
        self.calculator = calc

    def _build(self):
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        scroll.setStyleSheet("border: none; background: transparent;")

        inner = QWidget()
        inner.setStyleSheet("background: transparent;")
        lay = QVBoxLayout(inner)
        lay.setContentsMargins(18, 18, 18, 18)
        lay.setSpacing(12)

        lay.addWidget(SectionDivider("Параметры"))

        self.city_combo = AppComboBox()
        self.city_combo.addItems(["Москва", "Санкт-Петербург", "Екатеринбург", "Новосибирск", "Казань"])
        lay.addWidget(FormField("Город", self.city_combo))

        self.spec_combo = AppComboBox()
        self.spec_combo.addItems(["Кардиологи", "Терапевты", "Педиатры", "Аптеки"])
        lay.addWidget(FormField("Специализация", self.spec_combo))

        self.transport_combo = AppComboBox()
        self.transport_combo.addItems(["Автомобиль", "Общественный транспорт", "Пешком"])
        lay.addWidget(FormField("Транспорт", self.transport_combo))

        lay.addWidget(SectionDivider("Объём проекта"))

        self.total_visits_spin = AppSpinBox()
        self.total_visits_spin.setRange(10, 100000)
        self.total_visits_spin.setValue(500)
        self.total_visits_spin.setSuffix(" визитов")
        lay.addWidget(FormField("Всего визитов", self.total_visits_spin))

        self.visits_per_doctor_spin = AppSpinBox()
        self.visits_per_doctor_spin.setRange(1, 10)
        self.visits_per_doctor_spin.setValue(3)
        self.visits_per_doctor_spin.setSuffix(" визита")
        lay.addWidget(FormField("Визитов на врача", self.visits_per_doctor_spin))

        self.project_days_spin = AppSpinBox()
        self.project_days_spin.setRange(1, 365)
        self.project_days_spin.setValue(30)
        self.project_days_spin.setSuffix(" дней")
        lay.addWidget(FormField("Срок проекта", self.project_days_spin))

        lay.addSpacing(6)

        self.calculate_btn = AppButton("▶   Рассчитать проект", variant='primary')
        lay.addWidget(self.calculate_btn)

        hint = QLabel("Рассчитывает необходимое количество медпредов и варианты сроков реализации.")
        hint.setWordWrap(True)
        hint.setStyleSheet(f"color: {C['text3']}; font-size: 11px; line-height: 1.5;")
        lay.addWidget(hint)
        lay.addStretch()

        scroll.setWidget(inner)
        outer.addWidget(scroll)

    def refresh_theme(self):
        self.setStyleSheet(f"background-color: {C['panel']};")
        for w in self.findChildren(AppComboBox):    w.refresh_theme()
        for w in self.findChildren(AppSpinBox):     w.refresh_theme()
        for w in self.findChildren(AppButton):      w.refresh_theme()
        for w in self.findChildren(FormField):      w.refresh_theme()
        for w in self.findChildren(SectionDivider): w.refresh_theme()


# ─────────────────────────────────────────────────────────────────────────────
# ПАНЕЛЬ РЕЗУЛЬТАТОВ
# ─────────────────────────────────────────────────────────────────────────────

class ResultsPanel(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._build()

    def _ph(self, text):
        lbl = QLabel(text)
        lbl.setAlignment(Qt.AlignCenter)
        lbl.setStyleSheet(f"color: {C['text3']}; font-size: 15px; padding: 80px; background: transparent;")
        return lbl

    def _build(self):
        main_lay = QVBoxLayout(self)
        main_lay.setContentsMargins(0, 0, 0, 0)
        main_lay.setSpacing(0)

        # KPI-строка (скрыта до первого расчёта)
        self.metrics_widget = QWidget()
        self.metrics_widget.setFixedHeight(100)
        self.metrics_widget.setStyleSheet(
            f"background-color: {C['bg2']}; border-bottom: 1px solid {C['border']};")
        self.metrics_widget.setVisible(False)
        self.metric_cards_layout = QHBoxLayout(self.metrics_widget)
        self.metric_cards_layout.setContentsMargins(14, 10, 14, 10)
        self.metric_cards_layout.setSpacing(10)
        main_lay.addWidget(self.metrics_widget)

        # Стек страниц
        self.stack = QStackedWidget()
        self.stack.setStyleSheet(f"background-color: {C['bg']};")
        main_lay.addWidget(self.stack, stretch=1)

        # 0 — Графики
        self.graph_tab = QWidget()
        self.graph_tab.setStyleSheet(f"background-color: {C['bg']};")
        self.graph_layout = QVBoxLayout(self.graph_tab)
        self.graph_layout.setContentsMargins(12, 12, 12, 12)
        self.graph_layout.addWidget(self._ph("📊   Выполните расчёт для отображения графиков"))
        self.graph_canvas = None
        self.stack.addWidget(self.graph_tab)

        # 1 — Расписание
        self.schedule_tab = QWidget()
        self.schedule_tab.setStyleSheet(f"background-color: {C['bg']};")
        sl = QVBoxLayout(self.schedule_tab)
        sl.setContentsMargins(12, 12, 12, 12)
        self.schedule_table = AppTable(4, ['Время', 'Действие', 'Длительность', 'Тип'])
        sl.addWidget(self.schedule_table)
        self.stack.addWidget(self.schedule_tab)

        # 2 — Карта
        self.map_tab = QWidget()
        self.map_tab.setStyleSheet(f"background-color: {C['bg']};")
        ml = QVBoxLayout(self.map_tab)
        ml.setContentsMargins(0, 0, 0, 0)
        self.map_view = QWebEngineView()
        ml.addWidget(self.map_view)
        self.stack.addWidget(self.map_tab)

        # 3 — Рекомендации
        self.recommendations_tab = QWidget()
        self.recommendations_tab.setStyleSheet(f"background-color: {C['bg']};")
        rl = QVBoxLayout(self.recommendations_tab)
        rl.setContentsMargins(12, 12, 12, 12)
        self.recommendations_text = QTextEdit()
        self.recommendations_text.setReadOnly(True)
        self.recommendations_text.setStyleSheet(
            f"background-color: transparent; border: none; color: {C['text']};")
        rl.addWidget(self.recommendations_text)
        self.stack.addWidget(self.recommendations_tab)

        # 4 — Проект
        self.project_tab = QWidget()
        self.project_tab.setStyleSheet(f"background-color: {C['bg']};")
        po = QVBoxLayout(self.project_tab)
        po.setContentsMargins(0, 0, 0, 0)
        ps = QScrollArea(); ps.setWidgetResizable(True)
        pc = QWidget(); pc.setStyleSheet("background-color: transparent;")
        pcl = QVBoxLayout(pc)
        pcl.setContentsMargins(14, 14, 14, 14); pcl.setSpacing(12)

        self.project_report_text = QTextEdit()
        self.project_report_text.setReadOnly(True)
        self.project_report_text.setMaximumHeight(260)
        self.project_report_text.setStyleSheet(f"""
            QTextEdit {{
                background-color: {C['surface']}; border: 1px solid {C['border']};
                border-radius: 12px; color: {C['text']}; font-size: 13px; padding: 14px;
            }}
        """)
        pcl.addWidget(self.project_report_text)

        tbl_lbl = QLabel("ВАРИАНТЫ РЕАЛИЗАЦИИ ПРОЕКТА")
        tbl_lbl.setStyleSheet(
            f"color: {C['text3']}; font-size: 10px; font-weight: 700; letter-spacing: 1.5px; padding: 4px 0;")
        pcl.addWidget(tbl_lbl)

        self.project_scenarios_table = AppTable(8, [
            'Тип', 'Медпредов', 'Недель', 'Раб. дней',
            'Календ. дней', 'Исп. сроков%', 'Загрузка%', 'Рекомендация'
        ])
        self.project_scenarios_table.setMinimumHeight(320)
        pcl.addWidget(self.project_scenarios_table)

        pbr = QHBoxLayout()
        self.project_export_btn = AppButton("⬇   Экспорт в Excel", variant='ghost')
        self.project_export_btn.setEnabled(False)
        self.project_export_btn.setFixedWidth(180)
        pbr.addWidget(self.project_export_btn)
        pbr.addStretch()
        pcl.addLayout(pbr)

        ps.setWidget(pc)
        po.addWidget(ps)
        self.stack.addWidget(self.project_tab)

        # 5 — Монте-Карло
        self.monte_carlo_tab = QWidget()
        self.monte_carlo_tab.setStyleSheet(f"background-color: {C['bg']};")
        mc_main = QVBoxLayout(self.monte_carlo_tab)
        mc_main.setContentsMargins(0, 0, 0, 0); mc_main.setSpacing(0)

        mc_ctrl = QWidget()
        mc_ctrl.setFixedHeight(62)
        mc_ctrl.setStyleSheet(
            f"background-color: {C['bg2']}; border-bottom: 1px solid {C['border']};")
        mc_ctrl_lay = QHBoxLayout(mc_ctrl)
        mc_ctrl_lay.setContentsMargins(16, 0, 16, 0); mc_ctrl_lay.setSpacing(12)

        mc_iter_lbl = QLabel("Итераций:")
        mc_iter_lbl.setStyleSheet(f"color: {C['text2']}; font-size: 12px;")
        mc_ctrl_lay.addWidget(mc_iter_lbl)

        self.mc_iterations_spin = AppSpinBox()
        self.mc_iterations_spin.setRange(100, 10000)
        self.mc_iterations_spin.setValue(1000)
        self.mc_iterations_spin.setSuffix(" итераций")
        self.mc_iterations_spin.setFixedWidth(165)
        mc_ctrl_lay.addWidget(self.mc_iterations_spin)

        self.mc_run_btn = AppButton("▶   Запустить симуляцию", variant='primary')
        self.mc_run_btn.setEnabled(False)
        self.mc_run_btn.setFixedWidth(200)
        mc_ctrl_lay.addWidget(self.mc_run_btn)

        mc_hint = QLabel("Сначала выполните расчёт рабочего дня")
        mc_hint.setStyleSheet(f"color: {C['text3']}; font-size: 11px;")
        mc_ctrl_lay.addWidget(mc_hint)
        mc_ctrl_lay.addStretch()
        mc_main.addWidget(mc_ctrl)

        mc_scr = QScrollArea()
        mc_scr.setWidgetResizable(True)
        mc_scr.setStyleSheet("border: none; background: transparent;")
        mc_cnt = QWidget(); mc_cnt.setStyleSheet("background-color: transparent;")
        mc_cnt_lay = QVBoxLayout(mc_cnt)
        mc_cnt_lay.setContentsMargins(14, 14, 14, 14); mc_cnt_lay.setSpacing(14)

        self.mc_graph_widget = QWidget()
        self.mc_graph_widget.setStyleSheet("background-color: transparent;")
        self.mc_graph_layout = QVBoxLayout(self.mc_graph_widget)
        mc_cnt_lay.addWidget(self.mc_graph_widget)

        self.mc_stats_text = QTextEdit()
        self.mc_stats_text.setReadOnly(True)
        self.mc_stats_text.setMaximumHeight(220)
        self.mc_stats_text.setStyleSheet(f"""
            QTextEdit {{
                background-color: {C['surface']}; border: 1px solid {C['border']};
                border-radius: 12px; color: {C['text']}; padding: 14px;
            }}
        """)
        mc_cnt_lay.addWidget(self.mc_stats_text)
        mc_scr.setWidget(mc_cnt)
        mc_main.addWidget(mc_scr)

        self.stack.addWidget(self.monte_carlo_tab)

        # Нижний таббар
        self.tab_bar = BottomTabBar([
            ("📊", "Графики"),
            ("🗓", "Расписание"),
            ("🗺", "Карта"),
            ("💡", "Советы"),
            ("📋", "Проект"),
            ("🎲", "Монте-Карло"),
        ])
        self.tab_bar.tabChanged.connect(self.stack.setCurrentIndex)
        main_lay.addWidget(self.tab_bar)

    def set_current_tab(self, idx):
        self.stack.setCurrentIndex(idx)
        self.tab_bar.set_current(idx)

    def refresh_theme(self):
        self.metrics_widget.setStyleSheet(
            f"background-color: {C['bg2']}; border-bottom: 1px solid {C['border']};")
        self.stack.setStyleSheet(f"background-color: {C['bg']};")
        self.tab_bar.refresh_theme()
        for t in [self.graph_tab, self.schedule_tab, self.map_tab,
                  self.recommendations_tab, self.project_tab, self.monte_carlo_tab]:
            t.setStyleSheet(f"background-color: {C['bg']};")
        self.schedule_table.refresh_theme()
        self.project_scenarios_table.refresh_theme()
        self.recommendations_text.setStyleSheet(
            f"background-color: transparent; border: none; color: {C['text']};")
        self.project_report_text.setStyleSheet(f"""
            QTextEdit {{
                background-color: {C['surface']}; border: 1px solid {C['border']};
                border-radius: 12px; color: {C['text']}; font-size: 13px; padding: 14px;
            }}
        """)
        self.mc_stats_text.setStyleSheet(f"""
            QTextEdit {{
                background-color: {C['surface']}; border: 1px solid {C['border']};
                border-radius: 12px; color: {C['text']}; padding: 14px;
            }}
        """)
        for w in self.findChildren(AppButton):  w.refresh_theme()
        for w in self.findChildren(AppSpinBox): w.refresh_theme()


# ─────────────────────────────────────────────────────────────────────────────
# ДИАЛОГ ОБУЧЕНИЯ ML
# ─────────────────────────────────────────────────────────────────────────────

class TrainingDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Обучение ML-модели")
        self.setFixedSize(440, 185)
        self.setModal(True)
        self.setStyleSheet(f"background-color: {C['surface']}; color: {C['text']};")
        lay = QVBoxLayout(self)
        lay.setContentsMargins(28, 28, 28, 28)
        lay.setSpacing(14)

        info = QLabel("Обучение модели оптимизации. Это займёт 10–30 секунд...")
        info.setWordWrap(True)
        info.setAlignment(Qt.AlignCenter)
        info.setStyleSheet(f"font-size: 13px; color: {C['text2']};")
        lay.addWidget(info)

        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.progress_bar.setFixedHeight(8)
        self.progress_bar.setStyleSheet(f"""
            QProgressBar {{ border: none; border-radius: 4px;
                background-color: {C['surface2']}; text-align: center; color: transparent; }}
            QProgressBar::chunk {{ background-color: {C['accent']}; border-radius: 4px; }}
        """)
        lay.addWidget(self.progress_bar)

        self.status_label = QLabel("Подготовка данных...")
        self.status_label.setAlignment(Qt.AlignCenter)
        self.status_label.setStyleSheet(f"color: {C['text3']}; font-size: 12px;")
        lay.addWidget(self.status_label)

        row = QHBoxLayout()
        self.cancel_btn = AppButton("Отмена", variant='danger')
        self.cancel_btn.setFixedWidth(120)
        row.addStretch(); row.addWidget(self.cancel_btn)
        lay.addLayout(row)

    def update_progress(self, value, message=''):
        self.progress_bar.setValue(value)
        if message:
            self.status_label.setText(message)
        QApplication.processEvents()


# ─────────────────────────────────────────────────────────────────────────────
# ГЛАВНОЕ ОКНО
# ─────────────────────────────────────────────────────────────────────────────

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.calculator = MedicalRepCalculatorGUI()
        self.last_calculation_params = None
        self._build_ui()
        self._connect_signals()
        self._center()

    def _build_ui(self):
        self.setWindowTitle("ELOMER — Medical Rep Calculator")
        self.setMinimumSize(1280, 820)
        self.setStyleSheet(_stylesheet())

        central = QWidget()
        central.setStyleSheet(f"background-color: {C['bg']};")
        self.setCentralWidget(central)

        root = QHBoxLayout(central)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ── ЛЕВАЯ ПАНЕЛЬ ──────────────────────────────────────────────────────
        self.left_panel = QWidget()
        self.left_panel.setFixedWidth(285)
        self.left_panel.setStyleSheet(
            f"background-color: {C['panel']}; border-right: 1px solid {C['border']};")
        left_lay = QVBoxLayout(self.left_panel)
        left_lay.setContentsMargins(0, 0, 0, 0)
        left_lay.setSpacing(0)

        # Шапка с логотипом
        self.left_header = QWidget()
        self.left_header.setFixedHeight(70)
        self.left_header.setStyleSheet(
            f"background-color: {C['bg2']}; border-bottom: 1px solid {C['border']};")
        h_lay = QHBoxLayout(self.left_header)
        h_lay.setContentsMargins(16, 0, 16, 0)
        h_lay.setSpacing(11)

        logo = QLabel("E")
        logo.setFixedSize(38, 38)
        logo.setAlignment(Qt.AlignCenter)
        logo.setStyleSheet(f"""
            background: qlineargradient(x1:0,y1:0,x2:1,y2:1,
                stop:0 {C['accent']}, stop:1 {C['purple']});
            color: white; border-radius: 11px;
            font-size: 18px; font-weight: 800;
        """)
        h_lay.addWidget(logo)

        title_col = QVBoxLayout()
        title_col.setSpacing(1)
        self.title_lbl = QLabel("ELOMER")
        self.title_lbl.setStyleSheet(
            f"color: {C['text']}; font-size: 15px; font-weight: 800; "
            f"letter-spacing: 3px; background: transparent;")
        sub_lbl = QLabel("Med Rep Planner")
        sub_lbl.setStyleSheet(
            f"color: {C['text3']}; font-size: 10px; background: transparent; letter-spacing: 0.5px;")
        title_col.addWidget(self.title_lbl)
        title_col.addWidget(sub_lbl)
        h_lay.addLayout(title_col)
        h_lay.addStretch()

        self.ver_badge = QLabel("v1.0")
        self.ver_badge.setStyleSheet(f"""
            color: {C['text3']}; font-size: 10px; font-weight: 600;
            background: {C['surface2']}; border-radius: 6px;
            padding: 3px 8px; border: 1px solid {C['border2']};
        """)
        h_lay.addWidget(self.ver_badge)
        left_lay.addWidget(self.left_header)

        # Переключатель режима
        self.mode_bar = QWidget()
        self.mode_bar.setFixedHeight(50)
        self.mode_bar.setStyleSheet(
            f"background-color: {C['bg2']}; border-bottom: 1px solid {C['border']};")
        mb_lay = QHBoxLayout(self.mode_bar)
        mb_lay.setContentsMargins(14, 7, 14, 7)
        mb_lay.setSpacing(6)

        self.mode_daily_btn = QPushButton("Рабочий день")
        self.mode_proj_btn  = QPushButton("Проект")
        for btn in [self.mode_daily_btn, self.mode_proj_btn]:
            btn.setFixedHeight(32)
            btn.setCursor(Qt.PointingHandCursor)
        mb_lay.addWidget(self.mode_daily_btn)
        mb_lay.addWidget(self.mode_proj_btn)
        mb_lay.addStretch()
        left_lay.addWidget(self.mode_bar)

        self._apply_mode_btns(0)
        self.mode_daily_btn.clicked.connect(lambda: self._switch_mode(0))
        self.mode_proj_btn.clicked.connect(lambda: self._switch_mode(1))

        # Стек панелей ввода
        self.input_stack = QStackedWidget()
        self.input_stack.setStyleSheet("background: transparent;")
        self.daily_calc_panel   = DailyCalcPanel()
        self.project_calc_panel = ProjectCalcPanel()
        self.daily_calc_panel.set_calculator(self.calculator)
        self.project_calc_panel.set_calculator(self.calculator)
        self.input_stack.addWidget(self.daily_calc_panel)
        self.input_stack.addWidget(self.project_calc_panel)
        left_lay.addWidget(self.input_stack, stretch=1)

        # ML-кнопка
        self.ml_container = QWidget()
        self.ml_container.setFixedHeight(60)
        self.ml_container.setStyleSheet(
            f"background-color: {C['bg2']}; border-top: 1px solid {C['border']};")
        ml_lay = QHBoxLayout(self.ml_container)
        ml_lay.setContentsMargins(14, 10, 14, 10)
        self.train_btn = AppButton("🧠   Обучить ML-модель", variant='ghost')
        ml_lay.addWidget(self.train_btn)
        left_lay.addWidget(self.ml_container)

        # ── ПРАВАЯ ЧАСТЬ ──────────────────────────────────────────────────────
        right = QWidget()
        right.setStyleSheet(f"background-color: {C['bg']};")
        right_lay = QVBoxLayout(right)
        right_lay.setContentsMargins(0, 0, 0, 0)
        right_lay.setSpacing(0)

        self.right_header = QWidget()
        self.right_header.setFixedHeight(70)
        self.right_header.setStyleSheet(
            f"background-color: {C['bg2']}; border-bottom: 1px solid {C['border']};")
        rh_lay = QHBoxLayout(self.right_header)
        rh_lay.setContentsMargins(20, 0, 20, 0)
        rh_lay.setSpacing(14)

        self.results_title = QLabel("Результаты расчёта")
        self.results_title.setStyleSheet(
            f"color: {C['text']}; font-size: 17px; font-weight: 700; background: transparent;")
        rh_lay.addWidget(self.results_title)
        rh_lay.addStretch()

        self.status_dot = QLabel("●  Готов")
        self.status_dot.setStyleSheet(f"color: {C['success']}; font-size: 12px; font-weight: 600;")
        rh_lay.addWidget(self.status_dot)

        sep = QFrame()
        sep.setFrameShape(QFrame.VLine)
        sep.setFixedSize(1, 24)
        sep.setStyleSheet(f"background-color: {C['border2']};")
        rh_lay.addWidget(sep)

        self.theme_btn = QPushButton("☀  Светлая")
        self.theme_btn.setFixedSize(106, 34)
        self.theme_btn.setCursor(Qt.PointingHandCursor)
        self.theme_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {C['surface2']}; color: {C['text2']};
                border: 1px solid {C['border2']}; border-radius: 9px;
                font-size: 12px; font-weight: 600;
            }}
            QPushButton:hover {{ background-color: {C['accent']}; color: white; border-color: {C['accent']}; }}
        """)
        self.theme_btn.clicked.connect(self._toggle_theme)
        rh_lay.addWidget(self.theme_btn)
        right_lay.addWidget(self.right_header)

        self.results_panel = ResultsPanel()
        right_lay.addWidget(self.results_panel, stretch=1)

        root.addWidget(self.left_panel)
        root.addWidget(right, stretch=1)

        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_bar.showMessage("Готов к работе  ·  ELOMER v1.0")

        self._build_menu()
        self.results_panel.mc_run_btn.setEnabled(False)

    def _switch_mode(self, idx):
        self.input_stack.setCurrentIndex(idx)
        self._apply_mode_btns(idx)

    def _apply_mode_btns(self, active):
        for i, btn in enumerate([self.mode_daily_btn, self.mode_proj_btn]):
            if i == active:
                btn.setStyleSheet(f"""
                    QPushButton {{
                        background-color: {C['accent']}; color: white;
                        border: none; border-radius: 8px;
                        font-size: 12px; font-weight: 600; padding: 0 14px;
                    }}
                """)
            else:
                btn.setStyleSheet(f"""
                    QPushButton {{
                        background-color: transparent; color: {C['text2']};
                        border: 1px solid {C['border2']}; border-radius: 8px;
                        font-size: 12px; padding: 0 14px;
                    }}
                    QPushButton:hover {{ background-color: {C['surface2']}; color: {C['text']}; }}
                """)

    def _build_menu(self):
        mb = self.menuBar()
        fm = mb.addMenu("Файл")
        a = QAction("Новый расчёт", self); a.setShortcut("Ctrl+N")
        a.triggered.connect(self.reset_calculation); fm.addAction(a)
        fm.addSeparator()
        e = QAction("Экспорт результатов...", self); e.setShortcut("Ctrl+E")
        e.triggered.connect(self.export_results); fm.addAction(e)
        fm.addSeparator()
        q = QAction("Выход", self); q.setShortcut("Ctrl+Q")
        q.triggered.connect(self.close); fm.addAction(q)

        sm = mb.addMenu("Сервис")
        t = QAction("Обучить ML-модель", self); t.triggered.connect(self.train_model); sm.addAction(t)
        sm.addSeparator()
        s = QAction("Статистика по городам", self)
        s.triggered.connect(self.show_city_statistics); sm.addAction(s)

        hm = mb.addMenu("Справка")
        ab = QAction("О программе", self); ab.triggered.connect(self.show_about); hm.addAction(ab)

    def _connect_signals(self):
        self.daily_calc_panel.calculate_btn.clicked.connect(self.perform_calculation)
        self.daily_calc_panel.export_btn.clicked.connect(self.export_results)
        self.project_calc_panel.calculate_btn.clicked.connect(self.calculate_project)
        self.results_panel.mc_run_btn.clicked.connect(self.run_monte_carlo_simulation)
        self.results_panel.project_export_btn.clicked.connect(self.export_project_results)
        self.train_btn.clicked.connect(self.train_model)

    def _center(self):
        fg = self.frameGeometry()
        fg.moveCenter(self.screen().availableGeometry().center())
        self.move(fg.topLeft())

    def _toggle_theme(self):
        global _current_theme
        if _current_theme == 'dark':
            set_theme('light')
            self.theme_btn.setText("🌙  Тёмная")
        else:
            set_theme('dark')
            self.theme_btn.setText("☀  Светлая")
        self._apply_theme_to_all()

    def _apply_theme_to_all(self):
        self.setStyleSheet(_stylesheet())
        self.centralWidget().setStyleSheet(f"background-color: {C['bg']};")

        self.left_panel.setStyleSheet(
            f"background-color: {C['panel']}; border-right: 1px solid {C['border']};")
        self.left_header.setStyleSheet(
            f"background-color: {C['bg2']}; border-bottom: 1px solid {C['border']};")
        self.title_lbl.setStyleSheet(
            f"color: {C['text']}; font-size: 15px; font-weight: 800; "
            f"letter-spacing: 3px; background: transparent;")
        self.mode_bar.setStyleSheet(
            f"background-color: {C['bg2']}; border-bottom: 1px solid {C['border']};")
        self.ml_container.setStyleSheet(
            f"background-color: {C['bg2']}; border-top: 1px solid {C['border']};")
        self.right_header.setStyleSheet(
            f"background-color: {C['bg2']}; border-bottom: 1px solid {C['border']};")
        self.results_title.setStyleSheet(
            f"color: {C['text']}; font-size: 17px; font-weight: 700; background: transparent;")
        self.status_dot.setStyleSheet(f"color: {C['success']}; font-size: 12px; font-weight: 600;")
        self.ver_badge.setStyleSheet(f"""
            color: {C['text3']}; font-size: 10px; font-weight: 600;
            background: {C['surface2']}; border-radius: 6px;
            padding: 3px 8px; border: 1px solid {C['border2']};
        """)
        self.theme_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {C['surface2']}; color: {C['text2']};
                border: 1px solid {C['border2']}; border-radius: 9px;
                font-size: 12px; font-weight: 600;
            }}
            QPushButton:hover {{ background-color: {C['accent']}; color: white; border-color: {C['accent']}; }}
        """)
        self._apply_mode_btns(self.input_stack.currentIndex())
        self.train_btn.refresh_theme()
        self.daily_calc_panel.refresh_theme()
        self.project_calc_panel.refresh_theme()
        self.results_panel.refresh_theme()


    def perform_calculation(self):
        try:
            self.status_dot.setText("⏳  Расчёт...")
            self.status_dot.setStyleSheet(f"color: {C['warning']}; font-size: 12px; font-weight: 600;")
            QApplication.processEvents()
            city = self.daily_calc_panel.city_combo.currentText()
            spec = self.daily_calc_panel.spec_combo.currentText()
            transport = self.daily_calc_panel.transport_combo.currentText()
            num_visits = self.daily_calc_panel.visits_spin.value()

            if hasattr(self.calculator, 'unified_calculate_day_variable'):
                seed = int(time.time() * 1000) % 1000000
                result = self.calculator.unified_calculate_day_variable(
                    city, spec, num_visits, transport, seed)
                self.calculator.current_result = result
                if hasattr(self.calculator, 'calculate_daily_schedule'):
                    try:
                        detailed = self.calculator.calculate_daily_schedule(
                            city, spec, num_visits, transport)
                        result.update({
                            'schedule':  detailed.get('schedule', []),
                            'locations': detailed.get('locations', []),
                            'route':     detailed.get('route', [])
                        })
                    except Exception:
                        result['schedule'] = self._create_simple_schedule(result)
                        result['locations'] = self._generate_simple_locations(
                            city, spec, num_visits)
            else:
                result = self.calculator.calculate_daily_schedule(
                    city, spec, num_visits, transport)
                self.calculator.current_result = result

            self.update_results_display(result)
            self.last_calculation_params = {
                'city': city, 'specialization': spec,
                'transport': transport, 'num_visits': num_visits
            }
            self.results_panel.mc_run_btn.setEnabled(True)
            self.results_title.setText(f"{city}  ·  {spec}  ·  {num_visits} визитов")
            self.status_dot.setText("✓  Готов")
            self.status_dot.setStyleSheet(f"color: {C['success']}; font-size: 12px; font-weight: 600;")
            self.status_bar.showMessage(
                f"Расчёт завершён: {city}, {spec}, {num_visits} визитов")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка расчёта", str(e))

    def update_results_display(self, result):
        if not result:
            return
        try:
            self.update_graph_tab()
            self.update_schedule_table(result)
            if self.daily_calc_panel.show_map_checkbox.isChecked():
                self.update_map_tab()
            self.update_recommendations_tab()
            self.update_metrics_panel(result)
            self.results_panel.metrics_widget.setVisible(True)
            self.daily_calc_panel.export_btn.setEnabled(True)
            self.results_panel.set_current_tab(0)
        except Exception as e:
            QMessageBox.warning(self, "Ошибка отображения",
                                f"Не удалось отобразить результаты:\n{str(e)}")

    # ── ГРАФИКИ РАБОЧЕГО ДНЯ (оригинал: 4 графика) ──────────────────────────

    def update_graph_tab(self):
        while self.results_panel.graph_layout.count():
            item = self.results_panel.graph_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        try:
            from matplotlib.figure import Figure

            BG    = C['plt_bg']
            SURF  = C['plt_surf']
            TEXT  = C['plt_text']
            GRID  = C['plt_grid']
            BLUE  = C['plt_blue']
            GREEN = C['plt_green']
            ORANGE= C['plt_orange']
            RED   = C['plt_red']
            PURPLE= C['plt_purple']

            plt.rcParams.update({
                'axes.facecolor':   SURF,
                'figure.facecolor': BG,
                'axes.edgecolor':   GRID,
                'axes.labelcolor':  TEXT,
                'xtick.color':      TEXT,
                'ytick.color':      TEXT,
                'text.color':       TEXT,
                'grid.color':       GRID,
            })

            fig = Figure(figsize=(8.5, 7), dpi=100, facecolor=BG)
            result = self.calculator.current_result

            if not result:
                ax = fig.add_subplot(111)
                ax.set_facecolor(SURF)
                ax.text(0.5, 0.5, 'Нет данных\nВыполните расчет',
                        ha='center', va='center', transform=ax.transAxes,
                        fontsize=11, color=TEXT)
                ax.axis('off')
            else:
                gs = fig.add_gridspec(2, 2, hspace=0.25, wspace=0.2,
                                      left=0.09, right=0.96, top=0.93, bottom=0.07)

                # ── 1. КРУГОВАЯ: распределение времени ──────────────────────
                ax1 = fig.add_subplot(gs[0, 0])
                ax1.set_facecolor(SURF)
                time_data = {
                    'Перемещение': result.get('total_travel_time_min', 0),
                    'Визиты':      result.get('total_visit_time_min', 0)
                }
                if sum(time_data.values()) > 0:
                    wedges, texts, autotexts = ax1.pie(
                        time_data.values(),
                        labels=None,
                        autopct='%1.0f%%',
                        colors=[BLUE, GREEN],
                        startangle=90,
                        textprops={'fontsize': 9, 'fontweight': 'bold', 'color': 'white'},
                        wedgeprops={'linewidth': 0.5, 'edgecolor': BG},
                        pctdistance=0.75
                    )
                    ax1.legend(wedges, time_data.keys(),
                               loc='center left', bbox_to_anchor=(1.1, 0.5),
                               fontsize=8, frameon=False, labelcolor=TEXT)
                ax1.set_title('Распределение времени', fontsize=10,
                              fontweight='bold', pad=8, y=1, color=TEXT)

                # ── 2. SCATTER: время vs эффективность ──────────────────────
                ax2 = fig.add_subplot(gs[0, 1])
                ax2.set_facecolor(SURF)

                base_visits = result.get('num_visits', 8)
                visit_options, time_options, efficiency_options = [], [], []

                for delta in range(-3, 4):
                    test_visits = base_visits + delta
                    if test_visits < 1 or test_visits > 20:
                        continue
                    try:
                        base_visit_time  = result['total_visit_time_min']  / max(base_visits, 1)
                        base_travel_time = result['total_travel_time_min'] / max(base_visits, 1)
                        transport_type   = result.get('transport_type', 'Автомобиль')
                        transport_factors = {
                            'Автомобиль': 1.0, 'Общественный транспорт': 1.3, 'Пешком': 1.8
                        }
                        tf = transport_factors.get(transport_type, 1.0)

                        total_visit_time  = base_visit_time * test_visits
                        total_travel_time = base_travel_time * test_visits * tf
                        test_time  = total_visit_time + total_travel_time
                        test_hours = test_time / 60

                        if test_time > 0:
                            base_kpd = (total_visit_time / test_time) * 100
                            time_bonus = 15 if 7.0 <= test_hours <= 8.5 else \
                                         5  if 6.0 <= test_hours <= 9.0 else -10
                            visits_bonus = 10 if 6 <= test_visits <= 9 else \
                                           0  if 4 <= test_visits <= 11 else -15
                            efficiency = min(95, max(25, base_kpd + time_bonus + visits_bonus))
                        else:
                            efficiency = 50

                        visit_options.append(test_visits)
                        time_options.append(test_hours)
                        efficiency_options.append(efficiency)
                    except Exception:
                        continue

                if visit_options:
                    ax2b = ax2.twinx()

                    ax2.scatter(visit_options, time_options,
                                color=BLUE, s=60, alpha=0.7,
                                edgecolors='white', linewidth=1,
                                label='Общее время (ч)')
                    if len(visit_options) > 1:
                        z1 = np.polyfit(visit_options, time_options, 1)
                        x_trend = np.linspace(min(visit_options), max(visit_options), 100)
                        ax2.plot(x_trend, np.poly1d(z1)(x_trend),
                                 color=BLUE, linestyle='--', linewidth=1.5, alpha=0.5)

                    ax2b.scatter(visit_options, efficiency_options,
                                 color=GREEN, s=60, alpha=0.7,
                                 marker='s', edgecolors='white', linewidth=1,
                                 label='Эффективность (%)')
                    if len(visit_options) > 1:
                        z2 = np.polyfit(visit_options, efficiency_options, 2)
                        ax2b.plot(x_trend, np.poly1d(z2)(x_trend),
                                  color=GREEN, linestyle='--', linewidth=1.5, alpha=0.5)
                    ax2b.set_ylim([0, 110])
                    ax2b.set_ylabel('Эффективность (%)', fontsize=9, color=GREEN)
                    ax2b.tick_params(axis='y', labelcolor=GREEN)
                    ax2b.set_facecolor(SURF)

                    # Текущий план
                    if base_visits in visit_options:
                        ci = visit_options.index(base_visits)
                        ax2.scatter([base_visits], [time_options[ci]],
                                    color=RED, s=100, edgecolors='white', linewidth=2, zorder=5,
                                    label='Текущий план')
                        ax2b.scatter([base_visits], [efficiency_options[ci]],
                                     color=RED, s=100, marker='s',
                                     edgecolors='white', linewidth=2, zorder=5)
                        ax2.annotate(
                            f'{base_visits} визитов\n{time_options[ci]:.1f} ч',
                            xy=(base_visits, time_options[ci]),
                            xytext=(5, 15), textcoords='offset points',
                            fontsize=8, color=RED,
                            bbox=dict(boxstyle='round,pad=0.3',
                                      facecolor=SURF, alpha=0.9, edgecolor=GRID))

                    ax2.set_xlabel('Количество визитов', fontsize=9, color=TEXT)
                    ax2.set_ylabel('Общее время (часы)', fontsize=9, color=BLUE)
                    ax2.tick_params(axis='y', labelcolor=BLUE)
                    ax2.set_title('Анализ: время vs эффективность',
                                  fontsize=10, fontweight='bold', pad=8, color=TEXT)

                    lines1, labels1 = ax2.get_legend_handles_labels()
                    lines2, labels2 = ax2b.get_legend_handles_labels()
                    ax2.legend(lines1 + lines2, labels1 + labels2,
                               fontsize=7, loc='upper left',
                               framealpha=0.9,
                               facecolor=SURF, edgecolor=GRID, labelcolor=TEXT)
                    ax2.grid(True, alpha=0.2, linestyle=':', linewidth=0.5, color=GRID)
                    for spine in ax2.spines.values():
                        spine.set_edgecolor(GRID)

                # ── 3. BARH: расписание ──────────────────────────────────────
                ax3 = fig.add_subplot(gs[1, 0])
                ax3.set_facecolor(SURF)
                schedule = result.get('schedule', [])
                if schedule:
                    max_events = min(6, len(schedule))
                    schedule_to_show = schedule[:max_events]
                    activities, durations, bar_colors = [], [], []
                    for event in schedule_to_show:
                        act = event.get('activity', '')
                        if ('перемещение' in act.lower() or
                                event.get('type') == 'travel'):
                            activities.append('Перемещение')
                            bar_colors.append(BLUE)
                        else:
                            activities.append('Визит')
                            bar_colors.append(GREEN)
                        durations.append(event.get('duration_min', 0))

                    y_pos = range(len(activities))
                    bars_sched = ax3.barh(y_pos, durations,
                                          color=bar_colors, alpha=0.85, height=0.5)
                    ax3.set_yticks(y_pos)
                    ax3.set_yticklabels(activities, fontsize=8, color=TEXT)
                    ax3.set_xlabel('Мин.', fontsize=8, color=TEXT)
                    ax3.set_title('Расписание', fontsize=10,
                                  fontweight='bold', pad=8, color=TEXT)
                    ax3.invert_yaxis()
                    ax3.grid(True, axis='x', alpha=0.2, linestyle=':', linewidth=0.5, color=GRID)
                    ax3.set_axisbelow(True)
                    ax3.tick_params(colors=TEXT)
                    for spine in ax3.spines.values():
                        spine.set_edgecolor(GRID)

                    for bar, duration in zip(bars_sched, durations):
                        w = bar.get_width()
                        if w > max(durations) * 0.3:
                            ax3.text(w / 2, bar.get_y() + bar.get_height() / 2,
                                     f'{int(duration)}',
                                     ha='center', va='center',
                                     fontsize=7, color='white', fontweight='bold')
                        else:
                            ax3.text(w + max(durations) * 0.02,
                                     bar.get_y() + bar.get_height() / 2,
                                     f'{int(duration)}',
                                     va='center', fontsize=7, color=TEXT)
                else:
                    ax3.text(0.5, 0.5, 'Нет расписания',
                             ha='center', va='center', transform=ax3.transAxes,
                             fontsize=9, color=TEXT)
                    ax3.axis('off')

                # ── 4. BAR: эффективность транспорта ────────────────────────
                ax4 = fig.add_subplot(gs[1, 1])
                ax4.set_facecolor(SURF)
                ax4.set_position([0.55, 0.05, 0.42, 0.4])

                transports = ['Автомобиль', 'Общественный транспорт', 'Пешком']
                transports_display = ['АВТО', 'ТРАНСП', 'ХОДЬБА']
                colors_bar = [BLUE, PURPLE, GREEN]

                base_visit_time  = result['total_visit_time_min']  / max(result['num_visits'], 1)
                base_travel_time = result['total_travel_time_min'] / max(result['num_visits'], 1)
                speed_factors = {'Автомобиль': 1.0, 'Общественный транспорт': 1.4, 'Пешком': 2.0}

                times_per_visit, road_percentages = [], []
                for tr in transports:
                    adj_travel = base_travel_time * speed_factors[tr]
                    total_tpv  = base_visit_time + adj_travel
                    times_per_visit.append(total_tpv / 60)
                    road_percentages.append((adj_travel / total_tpv) * 100)

                bars = ax4.bar(transports_display, times_per_visit,
                               color=colors_bar, alpha=0.85,
                               width=0.6, edgecolor=BG, linewidth=2, zorder=3)

                for i, (bar, tv, rp) in enumerate(zip(bars, times_per_visit, road_percentages)):
                    h = bar.get_height()
                    if h > max(times_per_visit) * 0.4:
                        ax4.text(bar.get_x() + bar.get_width() / 2, h * 0.5,
                                 f'{tv:.1f} ч\n{rp:.0f}% дорога',
                                 ha='center', va='center',
                                 fontsize=9, fontweight='bold', color='white',
                                 linespacing=1.2)
                    else:
                        ax4.text(bar.get_x() + bar.get_width() / 2, h + 0.05,
                                 f'{tv:.1f} ч\n({rp:.0f}% дорога)',
                                 ha='center', va='bottom',
                                 fontsize=8, color=colors_bar[i], linespacing=1.1)

                # Выделение текущего транспорта
                cur_tr = result.get('transport_type', 'Автомобиль').lower()
                idx = (-1 if 'авто' not in cur_tr and 'машин' not in cur_tr
                       and 'общ' not in cur_tr and 'трансп' not in cur_tr
                       and 'пеш' not in cur_tr and 'ходь' not in cur_tr
                       else (0 if 'авто' in cur_tr or 'машин' in cur_tr
                             else 1 if 'общ' in cur_tr or 'трансп' in cur_tr
                             else 2))
                if idx >= 0:
                    ax4.text(idx, max(times_per_visit) * 1.12, '▼ ваш выбор',
                             ha='center', va='bottom', fontsize=9,
                             color=RED, fontweight='bold')
                    bars[idx].set_linewidth(4)
                    bars[idx].set_edgecolor(RED)

                if max(times_per_visit) > 1.8:
                    ax4.axhline(y=1.5, color=ORANGE, linestyle='--',
                                linewidth=1.5, alpha=0.6, zorder=2)
                    ax4.text(-0.3, 1.5 + 0.05, 'Цель: 1.5 ч',
                             fontsize=8, color=ORANGE, ha='left',
                             bbox=dict(boxstyle='round,pad=0.2',
                                       facecolor=SURF, alpha=0.8, edgecolor=GRID))

                ax4.set_ylabel('Часы на визит', fontsize=10, fontweight='bold',
                               labelpad=5, color=TEXT)
                ax4.set_xlabel('Вид транспорта', fontsize=10, fontweight='bold',
                               labelpad=8, color=TEXT)
                ax4.set_title('Эффективность транспорта', fontsize=10,
                              fontweight='bold', pad=7, color=TEXT)
                y_max = max(times_per_visit) * 1.25
                ax4.set_ylim([0, y_max])
                ax4.set_yticks(np.arange(0, y_max + 0.5, 0.5))
                ax4.tick_params(axis='both', labelsize=9, colors=TEXT)
                ax4.grid(True, axis='y', alpha=0.2, linestyle=':', linewidth=0.5,
                         color=GRID, zorder=1)
                ax4.set_axisbelow(True)
                ax4.spines['top'].set_visible(False)
                ax4.spines['right'].set_visible(False)
                for sp in ['bottom', 'left']:
                    ax4.spines[sp].set_edgecolor(GRID)

            fig.tight_layout(rect=[0, 0.03, 1, 0.97])
            canvas = FigureCanvas(fig)
            self.results_panel.graph_layout.addWidget(canvas)

        except Exception as e:
            err = QLabel(f"Ошибка построения графика: {str(e)[:80]}")
            err.setAlignment(Qt.AlignCenter)
            err.setStyleSheet(f"color: {C['danger']}; font-size: 12px; padding: 20px;")
            self.results_panel.graph_layout.addWidget(err)

    # ── Расписание ────────────────────────────────────────────────────────────

    def update_schedule_table(self, result):
        table = self.results_panel.schedule_table
        schedule = result.get('schedule', [])
        table.setRowCount(len(schedule))
        type_labels = {'visit': 'Визит', 'travel': 'Переезд', 'admin': 'Администрация'}
        type_colors = {'visit': C['success'], 'travel': C['accent'], 'admin': C['warning']}

        for i, event in enumerate(schedule):
            etype = event.get('type', 'visit')
            table.set_cell(i, 0, event.get('time', ''))
            table.set_cell(i, 1, event.get('activity', ''))
            table.set_cell(i, 2, f"{event.get('duration_min', '')} мин",
                           align=Qt.AlignCenter)

            item = QTableWidgetItem(type_labels.get(etype, etype.capitalize()))
            item.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
            item.setForeground(QColor(type_colors.get(etype, C['text2'])))
            f = item.font(); f.setBold(True); item.setFont(f)
            table.setItem(i, 3, item)
        table.resizeColumnsToContents()

    def update_map_tab(self):
        html = self.calculator.create_interactive_map()
        if html:
            self.results_panel.map_view.setHtml(html)

    def update_recommendations_tab(self):
        recs = self.calculator.get_recommendations()
        if not recs:
            return

        is_dark = _current_theme == 'dark'
        tag_names = {'warning': 'Внимание', 'info': 'Информация', 'success': 'Рекомендация'}
        border_colors = {
            'warning': C['warning'], 'info': C['accent'], 'success': C['success']
        }
        if is_dark:
            bg_colors = {'warning': '#2A2410', 'info': '#1A1830', 'success': '#0F2410'}
        else:
            bg_colors = {'warning': '#FFF8EC', 'info': '#EBF5FB', 'success': '#EAFAF1'}

        html = f"""<html><head><style>
        body {{ font-family:'Segoe UI',Arial; background:transparent;
               color:{C['text']}; margin:0; padding:0; }}
        .card {{ padding:12px 16px; margin-bottom:10px; border-radius:8px;
                border-left:3px solid; font-size:13px; }}
        .tag {{ font-size:10px; font-weight:700; letter-spacing:1px; opacity:.7;
               text-transform:uppercase; margin-bottom:4px; }}
        </style></head><body>"""
        for rec in recs:
            t = rec.get('type', 'info')
            bc = border_colors.get(t, C['accent'])
            bgc = bg_colors.get(t, bg_colors['info'])
            tag_text = tag_names.get(t, '')
            html += (f'<div class="card" style="border-color:{bc};background-color:{bgc};">'
                     f'<div class="tag">{tag_text}</div>{rec.get("text","")}</div>')
        html += "</body></html>"
        self.results_panel.recommendations_text.setHtml(html)

    def update_metrics_panel(self, result):
        if not result:
            return
        try:
            lay = self.results_panel.metric_cards_layout
            while lay.count():
                item = lay.takeAt(0)
                if item.widget():
                    item.widget().deleteLater()

            tv = result.get('total_visit_time_min', 0)
            tt = result.get('total_work_time_min', 1)
            kpd = min(100, max(0, (tv / tt) * 100))
            result['efficiency_score'] = kpd

            for title, value, unit, color in [
                ("Общее время",   f"{result.get('total_work_hours', 0):.1f}", "часов", C['accent']),
                ("Визитов",       str(result.get('num_visits', 0)),            "",      C['success']),
                ("Дистанция",     f"{result.get('total_distance_km', 0):.1f}", "км",    C['danger']),
                ("Эффективность", f"{kpd:.1f}",                                "%",     C['warning']),
                ("Загрузка",      f"{result.get('work_day_utilization', 0):.1f}", "%", C['purple']),
            ]:
                card = MetricCard(title, value, unit, color)
                card.setMinimumWidth(110)
                lay.addWidget(card)
            lay.addStretch()
            self.results_panel.metrics_widget.setVisible(True)
        except Exception as e:
            print(f"Ошибка метрик: {e}")

    # ── Проект ───────────────────────────────────────────────────────────────

    def calculate_project(self):
        try:
            self.status_dot.setText("⏳  Расчёт...")
            self.status_dot.setStyleSheet(f"color: {C['warning']}; font-size: 12px; font-weight: 600;")
            QApplication.processEvents()
            city      = self.project_calc_panel.city_combo.currentText()
            spec      = self.project_calc_panel.spec_combo.currentText()
            transport = self.project_calc_panel.transport_combo.currentText()
            total_visits    = self.project_calc_panel.total_visits_spin.value()
            vpd             = self.project_calc_panel.visits_per_doctor_spin.value()
            proj_days       = self.project_calc_panel.project_days_spin.value()

            result = self.calculator.calculate_city_load(
                city, spec, transport, total_visits, vpd, proj_days)
            self.calculator.current_project_result = result
            self._show_project_results(result)
            self.results_panel.project_export_btn.setEnabled(True)
            self.results_title.setText(f"Проект: {city}  ·  {spec}")
            self.status_bar.showMessage(f"Проектный расчёт завершён: {city}, {spec}")
            self.status_dot.setText("✓  Готов")
            self.status_dot.setStyleSheet(f"color: {C['success']}; font-size: 12px; font-weight: 600;")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка расчёта проекта", f"Произошла ошибка:\n{str(e)}")

    def _show_project_results(self, result):
        self.results_panel.set_current_tab(4)
        self.results_panel.project_report_text.setHtml(
            self._format_project_report(result))
        self._fill_project_scenarios_table(result)
        self.results_panel.metrics_widget.setVisible(False)

    def _format_project_report(self, result):
        if 'error' in result:
            return f"<span style='color:{C['danger']}'>{result['error']}</span>"
        calc = result['calculations']
        ip   = result['input_params']
        TC   = C['text']
        TC2  = C['text2']

        def row(label, value, color=None):
            vs = (f"color:{color};font-weight:700;font-size:15px;"
                  if color else f"color:{TC};")
            return (f"<tr><td style='color:{TC2};padding:5px 0;'>{label}</td>"
                    f"<td style='{vs}padding:5px 12px;'>{value}</td></tr>")

        return f"""<html><head><style>
        body{{font-family:'Segoe UI',Arial;background:transparent;
             color:{TC};margin:0;padding:0;}}
        h3{{color:{TC2};font-size:10px;font-weight:700;letter-spacing:1.2px;
            text-transform:uppercase;margin:14px 0 5px;}}
        table{{width:100%;border-collapse:collapse;}}
        </style></head><body>
        <h3>Исходные параметры</h3><table>
        {row("Город",           result['city'])}
        {row("Специализация",   result['specialization'])}
        {row("Транспорт",       result['transport_type'])}
        {row("Всего визитов",   f"{ip['total_visits_needed']:,}")}
        {row("Визитов на врача",ip['visits_per_doctor'])}
        {row("Срок проекта",    f"{ip['project_calendar_days']} календ. дней")}
        {row("Рабочих дней/нед.",ip['work_days_per_week'])}
        </table>
        <h3>Результаты расчёта</h3><table>
        {row("Уникальных врачей",       f"{calc['unique_doctors_needed']:,}")}
        {row("Время на одного врача",   f"{calc['time_per_doctor_hours']} ч")}
        {row("Общее время работы",      f"{calc['total_time_all_doctors_hours']:,} ч")}
        {row("Мин. медпредов",          calc['min_reps_needed'],         C['danger'])}
        {row("Оптимально медпредов",    calc.get('optimal_reps_needed',
                                                  calc['min_reps_needed']), C['success'])}
        {row("Напряжённость",
             f"{calc.get('project_status','—')} ({calc.get('project_intensity',0)}%)")}
        </table>
        <h3>Пример рабочего дня</h3><table>
        {row("Визитов в день",  result['standard_day_example']['visits_per_day'])}
        {row("Время работы",    f"{result['standard_day_example']['work_hours']:.1f} ч")}
        {row("Расстояние",      f"{result['standard_day_example']['distance_km']:.1f} км")}
        </table></body></html>"""

    def _fill_project_scenarios_table(self, result):
        table = self.results_panel.project_scenarios_table
        if 'scenarios' not in result or not result['scenarios']:
            table.setRowCount(0)
            return
        scenarios = result['scenarios'][:15]
        table.setRowCount(len(scenarios))

        for i, sc in enumerate(scenarios):
            is_opt = sc.get('is_optimal', False)
            is_min = sc.get('is_minimal', False)

            if is_opt:
                table.set_cell(i, 0, "ОПТИМАЛЬНЫЙ", fg_color=C['success'],
                               bold=True, align=Qt.AlignCenter)
            elif is_min:
                table.set_cell(i, 0, "МИНИМАЛЬНЫЙ", fg_color=C['danger'],
                               bold=True, align=Qt.AlignCenter)
            else:
                ru = sc.get('rep_utilization', 50)
                label = ("Высокая нагрузка" if ru > 90 else
                         "Норма"            if ru > 60 else "Низкая нагрузка")
                table.set_cell(i, 0, label, align=Qt.AlignCenter)

            table.set_cell(i, 1, str(sc['reps_count']), align=Qt.AlignCenter)
            table.set_cell(i, 2, str(sc['weeks']), align=Qt.AlignCenter)
            table.set_cell(i, 3, str(int(sc['work_days'])), align=Qt.AlignCenter)
            table.set_cell(i, 4, str(int(sc['calendar_days'])), align=Qt.AlignCenter)

            tu = sc['time_utilization']
            table.set_cell(i, 5, f"{tu}%",
                           fg_color=(C['danger']  if tu > 90 else
                                     C['warning'] if tu > 70 else C['success']),
                           align=Qt.AlignCenter, bold=(tu > 90))

            ru = sc['rep_utilization']
            table.set_cell(i, 6, f"{ru}%",
                           fg_color=(C['danger']  if ru > 90 else
                                     C['warning'] if ru > 70 else
                                     C['success'] if ru > 60 else C['purple']),
                           align=Qt.AlignCenter, bold=(ru > 90))

            table.set_cell(i, 7, sc.get('recommendation', ''))

        table.resizeColumnsToContents()
        table.horizontalHeader().setSectionResizeMode(7, QHeaderView.Stretch)

    # ── Монте-Карло ──────────────────────────────────────────────────────────

    def run_monte_carlo_simulation(self):
        try:
            if not self.last_calculation_params:
                QMessageBox.information(self, "Сначала выполните расчёт",
                    "Выполните расчёт рабочего дня, затем запустите симуляцию.")
                return
            p     = self.last_calculation_params
            iters = self.results_panel.mc_iterations_spin.value()
            mc    = self.calculator.monte_carlo_daily_simulation(
                p['city'], p['specialization'], p['num_visits'], p['transport'], iters)
            self.calculator.current_mc_results = mc
            self.update_monte_carlo_graphs(mc)
            self.update_mc_statistics(mc)
            self.results_panel.set_current_tab(5)
            self.status_bar.showMessage(f"Монте-Карло завершено ({iters} итераций)")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка симуляции", str(e))

    def update_monte_carlo_graphs(self, mc_results):
        while self.results_panel.mc_graph_layout.count():
            item = self.results_panel.mc_graph_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        try:
            from matplotlib.figure import Figure

            BG    = C['plt_bg']
            SURF  = C['plt_surf']
            TEXT  = C['plt_text']
            GRID  = C['plt_grid']
            BLUE  = C['plt_blue']
            GREEN = C['plt_green']
            ORANGE= C['plt_orange']
            RED   = C['plt_red']
            PURPLE= C['plt_purple']

            plt.rcParams.update({
                'axes.facecolor':   SURF,
                'figure.facecolor': BG,
                'axes.edgecolor':   GRID,
                'axes.labelcolor':  TEXT,
                'xtick.color':      TEXT,
                'ytick.color':      TEXT,
                'text.color':       TEXT,
                'grid.color':       GRID,
            })

            if not mc_results or 'raw_results' not in mc_results:
                err = QLabel("Нет данных для построения графиков")
                err.setAlignment(Qt.AlignCenter)
                err.setStyleSheet(f"color: {C['text3']}; padding: 40px;")
                self.results_panel.mc_graph_layout.addWidget(err)
                return

            raw = mc_results['raw_results']
            fig = Figure(figsize=(10, 16), dpi=100, facecolor=BG)
            fig.subplots_adjust(left=0.12, right=0.95, bottom=0.07, top=0.94,
                                hspace=0.4, wspace=0.3)

            # ── 1. ГИСТОГРАММА времени работы (с KDE) ────────────────────────
            ax1 = fig.add_subplot(311)
            ax1.set_facecolor(SURF)
            if 'total_hours' in raw and len(raw['total_hours']) > 0:
                total_hours_filtered = [x for x in raw['total_hours']
                                        if x is not None and not np.isnan(x)]
                if total_hours_filtered:
                    n, bins, patches = ax1.hist(total_hours_filtered, bins=30,
                                                alpha=0.7, color=BLUE,
                                                edgecolor=BG, linewidth=1)
                    # KDE
                    try:
                        from scipy.stats import gaussian_kde
                        kde = gaussian_kde(total_hours_filtered)
                        x_range = np.linspace(min(total_hours_filtered),
                                               max(total_hours_filtered), 150)
                        density = kde(x_range)
                        ax1_twin = ax1.twinx()
                        ax1_twin.plot(
                            x_range,
                            density * len(total_hours_filtered) * (bins[1] - bins[0]),
                            color=RED, linewidth=2, alpha=0.8, label='Плотность')
                        ax1_twin.set_ylabel('Плотность', fontsize=9, color=TEXT)
                        ax1_twin.tick_params(axis='y', labelsize=8, colors=TEXT)
                        ax1_twin.set_facecolor(SURF)
                    except Exception:
                        pass

                    mean_val   = np.mean(total_hours_filtered)
                    median_val = np.median(total_hours_filtered)
                    ax1.axvline(mean_val,   color=RED,    linestyle='--',
                                linewidth=2.5, label=f'Среднее: {mean_val:.1f} ч')
                    ax1.axvline(median_val, color=ORANGE, linestyle=':',
                                linewidth=2,   label=f'Медиана: {median_val:.1f} ч')
                    ax1.axvline(8, color=GREEN, linestyle='-',
                                linewidth=1.5, alpha=0.8, label='Норма (8 ч)')

                    ax1.set_xlabel('Общее время работы (часы)', fontsize=11,
                                   fontweight='bold', color=TEXT)
                    ax1.set_ylabel('Количество дней', fontsize=11,
                                   fontweight='bold', color=TEXT)
                    ax1.set_title('РАСПРЕДЕЛЕНИЕ ВРЕМЕНИ РАБОТЫ',
                                  fontsize=13, fontweight='bold', pad=12, color=TEXT)

                    handles1, labels1 = ax1.get_legend_handles_labels()
                    if 'ax1_twin' in dir():
                        handles2, labels2 = ax1_twin.get_legend_handles_labels()
                        ax1.legend(handles1 + handles2, labels1 + labels2,
                                   fontsize=9, loc='upper right', framealpha=0.9,
                                   facecolor=SURF, edgecolor=GRID, labelcolor=TEXT)
                    else:
                        ax1.legend(fontsize=9, loc='upper right', framealpha=0.9,
                                   facecolor=SURF, edgecolor=GRID, labelcolor=TEXT)

                    ax1.grid(True, alpha=0.2, linestyle='--', color=GRID)
                    ax1.tick_params(axis='both', labelsize=9, colors=TEXT)

                    stats_text = (
                        f"Всего симуляций: {len(total_hours_filtered):,}\n"
                        f"Среднее: {mean_val:.1f} ч\n"
                        f"Медиана: {median_val:.1f} ч\n"
                        f"Стд. отклонение: {np.std(total_hours_filtered):.1f} ч")
                    ax1.text(0.98, 0.98, stats_text,
                             transform=ax1.transAxes,
                             fontsize=9, verticalalignment='top',
                             horizontalalignment='right',
                             bbox=dict(boxstyle='round', facecolor=SURF,
                                       alpha=0.9, edgecolor=GRID), color=TEXT)
                    for sp in ax1.spines.values():
                        sp.set_edgecolor(GRID)

            # ── 2. ГИСТОГРАММА эффективности с цветными зонами ───────────────
            ax2 = fig.add_subplot(312)
            ax2.set_facecolor(SURF)
            if 'efficiency' in raw and len(raw['efficiency']) > 0:
                eff_filtered = [x for x in raw['efficiency']
                                if x is not None and not np.isnan(x)]
                if eff_filtered:
                    n_bins = min(20, len(eff_filtered) // 10)
                    ax2.hist(eff_filtered, bins=n_bins, alpha=0.7,
                             color=GREEN, edgecolor=BG, linewidth=1)

                    mean_eff = np.mean(eff_filtered)
                    ax2.axvline(mean_eff, color=RED, linestyle='--',
                                linewidth=2.5, label=f'Среднее: {mean_eff:.1f}%')

                    ax2.axvspan(0,  60,  alpha=0.1, color=RED,    label='Низкая (<60%)')
                    ax2.axvspan(60, 80,  alpha=0.1, color=ORANGE, label='Средняя (60-80%)')
                    ax2.axvspan(80, 100, alpha=0.1, color=GREEN,  label='Высокая (>80%)')

                    ax2.set_xlabel('КПД работы (%)', fontsize=11, fontweight='bold', color=TEXT)
                    ax2.set_ylabel('Количество дней', fontsize=11, fontweight='bold', color=TEXT)
                    ax2.set_title('РАСПРЕДЕЛЕНИЕ ЭФФЕКТИВНОСТИ',
                                  fontsize=13, fontweight='bold', pad=12, color=TEXT)
                    ax2.legend(fontsize=9, loc='upper left', framealpha=0.9,
                               facecolor=SURF, edgecolor=GRID, labelcolor=TEXT)
                    ax2.grid(True, alpha=0.2, linestyle='--', color=GRID)
                    ax2.tick_params(axis='both', labelsize=9, colors=TEXT)
                    for sp in ax2.spines.values():
                        sp.set_edgecolor(GRID)

            # ── 3. SCATTER: время работы vs визиты (с трендом и зонами) ──────
            ax3 = fig.add_subplot(313)
            ax3.set_facecolor(SURF)
            has_hours  = 'total_hours'       in raw and len(raw['total_hours']) > 0
            has_visits = 'successful_visits' in raw and len(raw['successful_visits']) > 0
            if has_hours and has_visits:
                hours  = raw['total_hours']
                visits = raw['successful_visits']
                valid  = [(h, v) for h, v in zip(hours, visits)
                          if h is not None and v is not None
                          and not np.isnan(h) and not np.isnan(v)]
                if valid and len(valid) >= 10:
                    hours_v, visits_v = zip(*valid)
                    ax3.scatter(hours_v, visits_v, alpha=0.6, s=40,
                                color=PURPLE, edgecolors='white', linewidth=0.5)
                    if len(hours_v) > 1:
                        z = np.polyfit(hours_v, visits_v, 1)
                        x_trend = np.linspace(min(hours_v), max(hours_v), 100)
                        ax3.plot(x_trend, np.poly1d(z)(x_trend),
                                 color=RED, linewidth=2, linestyle='--',
                                 alpha=0.9, label='Линия тренда')
                    ax3.axvspan(5, 8,  alpha=0.15, color=GREEN)
                    ax3.axhspan(6, 12, alpha=0.15, color=BLUE)

                    ax3.set_xlabel('Общее время работы (часы)', fontsize=11,
                                   fontweight='bold', color=TEXT)
                    ax3.set_ylabel('Успешные визиты', fontsize=11,
                                   fontweight='bold', color=TEXT)
                    ax3.set_title('ЗАВИСИМОСТЬ: ВРЕМЯ РАБОТЫ vs КОЛИЧЕСТВО ВИЗИТОВ',
                                  fontsize=13, fontweight='bold', pad=12, color=TEXT)
                    ax3.legend(fontsize=9, loc='upper left', framealpha=0.9,
                               facecolor=SURF, edgecolor=GRID, labelcolor=TEXT)
                    ax3.grid(True, alpha=0.2, linestyle='--', color=GRID)
                    ax3.tick_params(axis='both', labelsize=9, colors=TEXT)
                    for sp in ax3.spines.values():
                        sp.set_edgecolor(GRID)

            if 'input_params' in mc_results:
                pr = mc_results['input_params']
                fig.suptitle(
                    f"АНАЛИЗ МОНТЕ-КАРЛО | {pr.get('city','?')} | "
                    f"{pr.get('iterations',0):,} итераций",
                    fontsize=14, fontweight='bold', y=0.98, color=TEXT)

            canvas = FigureCanvas(fig)
            canvas.setMinimumSize(900, 1400)
            self.results_panel.mc_graph_layout.addWidget(canvas)

        except Exception as e:
            err = QLabel(f"Ошибка построения графиков: {str(e)}")
            err.setAlignment(Qt.AlignCenter)
            err.setStyleSheet(f"color: {C['danger']}; padding: 20px;")
            self.results_panel.mc_graph_layout.addWidget(err)

    def update_mc_statistics(self, mc_results):
        if not mc_results or 'statistics' not in mc_results:
            return
        stats = mc_results['statistics']
        ip    = mc_results['input_params']
        TC  = C['text']
        TC2 = C['text2']

        def sr(lbl, val):
            return (f"<tr><td style='color:{TC2};padding:4px 0;'>{lbl}</td>"
                    f"<td style='color:{TC};font-weight:600;padding:4px 12px;'>{val}</td></tr>")

        html = f"""<html><head><style>
        body{{font-family:'Segoe UI';background:transparent;color:{TC};margin:0;padding:0;}}
        h3{{color:{TC2};font-size:10px;font-weight:700;letter-spacing:1.2px;
            text-transform:uppercase;margin:10px 0 5px;}}
        table{{border-collapse:collapse;width:100%;}}
        </style></head><body>
        <h3>Параметры симуляции</h3><table>
        {sr("Город",           ip.get('city','?'))}
        {sr("Специализация",   ip.get('specialization','?'))}
        {sr("Визитов",         ip.get('num_visits','?'))}
        {sr("Итераций",        f"{ip.get('iterations',0):,}")}
        </table>"""
        if 'total_hours' in stats:
            th = stats['total_hours']
            html += f"""<h3>Время работы</h3><table>
            {sr("Среднее",          f"{th['mean']:.1f} ч")}
            {sr("Медиана",          f"{th['median']:.1f} ч")}
            {sr("Стд. отклонение",  f"{th['std']:.1f} ч")}
            </table>"""
        if 'overload_probability' in stats:
            op    = stats['overload_probability']
            color = C['danger'] if op.get('value', 0) > 20 else C['warning']
            html += (f"<h3>Анализ рисков</h3>"
                     f"<p style='color:{color};font-weight:600;'>"
                     f"{op.get('description','')}</p>")
        html += "</body></html>"
        self.results_panel.mc_stats_text.setHtml(html)

    # ── ML ───────────────────────────────────────────────────────────────────

    def train_model(self):
        if hasattr(self, '_training_thread') and self._training_thread.isRunning():
            QMessageBox.information(self, "Обучение", "Обучение уже выполняется...")
            return
        self.training_dialog = TrainingDialog(self)
        self.training_dialog.cancel_btn.clicked.connect(self.cancel_training)
        from threading import Thread
        self._training_thread = Thread(target=self._train_model_thread)
        self._training_thread.start()
        self.training_dialog.exec()

    def _train_model_thread(self):
        try:
            from PySide6.QtCore import QTimer
            for i, step in enumerate(["Подготовка данных", "Создание модели",
                                       "Обучение", "Сохранение"]):
                p = (i + 1) * 25
                QTimer.singleShot(0, lambda pv=p, sv=step:
                    self.training_dialog.update_progress(pv, f"{sv}... {pv}%"))
                time.sleep(0.3)
            model, score = self.calculator.train_optimization_model(
                lambda pv: QTimer.singleShot(0, lambda: self.training_dialog.update_progress(
                    pv, f"Обучение... {pv}%")))
            QTimer.singleShot(500, self.training_dialog.accept)
            QTimer.singleShot(600, lambda: QMessageBox.information(
                self, "Обучение завершено",
                "Модель готова к работе!\n\nПри расчётах с включённой опцией 'Использовать ML'\n"
                "система будет рекомендовать оптимальное количество визитов."))
        except Exception as e:
            print(f"Ошибка в потоке: {e}")
            from PySide6.QtCore import QTimer
            QTimer.singleShot(0, self.training_dialog.reject)
            QTimer.singleShot(0, lambda: QMessageBox.warning(
                self, "Обучение не удалось",
                "Модель не обучена, но приложение продолжит работу.\n"
                "Используйте стандартный расчёт без ML-оптимизации."))

    def cancel_training(self):
        if hasattr(self, '_training_thread') and self._training_thread.isRunning():
            self.training_dialog.reject()
            QMessageBox.information(self, "Обучение", "Обучение прервано пользователем")

    # ── Экспорт ──────────────────────────────────────────────────────────────

    def export_results(self):
        if not self.calculator.current_result:
            QMessageBox.warning(self, "Нет данных", "Сначала выполните расчёт")
            return
        try:
            fn, _ = QFileDialog.getSaveFileName(
                self, "Экспорт результатов", "",
                "Excel файлы (*.xlsx);;Все файлы (*)")
            if fn:
                if not fn.endswith('.xlsx'):
                    fn += '.xlsx'
                self._export_daily_to_excel(fn)
                QMessageBox.information(self, "Успех", f"Результаты экспортированы в {fn}")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка экспорта", str(e))

    def _export_daily_to_excel(self, fn):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        result = self.calculator.current_result
        wb = Workbook()
        ws1 = wb.active; ws1.title = "Сводка"
        ws1.merge_cells('A1:D1')
        tc = ws1['A1']
        tc.value = f"Результаты расчёта: {result['city']} — {result['specialization']}"
        tc.font = Font(bold=True, size=14)
        tc.alignment = Alignment(horizontal='center')
        hf = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        metrics = [
            ["Параметр", "Значение", "Единица"],
            ["Город", result['city'], ""],
            ["Специализация", result['specialization'], ""],
            ["Транспорт", result['transport_type'], ""],
            ["Количество визитов", result['num_visits'], ""],
            ["Общее время работы", result['total_work_hours'], "часов"],
            ["Общее расстояние", result['total_distance_km'], "км"],
            ["Время в пути", result['total_travel_time_min'], "минут"],
            ["Время на визиты", result['total_visit_time_min'], "минут"],
            ["Загрузка рабочего дня", result['work_day_utilization'], "%"],
            ["Эффективность", result.get('efficiency_score', 0), "%"],
        ]
        for i, row in enumerate(metrics, start=3):
            for j, val in enumerate(row):
                cell = ws1.cell(row=i, column=j + 1, value=val)
                if i == 3:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = hf
        ws2 = wb.create_sheet("Расписание")
        ws2.append(["Время", "Действие", "Длительность (мин)", "Тип"])
        for ev in result.get('schedule', []):
            ws2.append([ev['time'], ev['activity'], ev['duration_min'],
                        ev.get('type', '').capitalize()])
        for cell in ws2[1]:
            cell.font = Font(bold=True, color="FFFFFF"); cell.fill = hf
        for ws in [ws1, ws2]:
            for col in ws.columns:
                ml = max((len(str(c.value)) for c in col if c.value), default=0)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 2, 50)
        wb.save(fn)

    def export_project_results(self):
        res = getattr(self.calculator, 'current_project_result', None)
        if not res:
            QMessageBox.warning(self, "Нет данных", "Сначала выполните расчёт проекта")
            return
        try:
            fn, _ = QFileDialog.getSaveFileName(
                self, "Экспорт результатов проекта", "",
                "Excel файлы (*.xlsx);;Все файлы (*)")
            if fn:
                if not fn.endswith('.xlsx'):
                    fn += '.xlsx'
                self._export_project_to_excel(fn, res)
                QMessageBox.information(self, "Успех", f"Проект экспортирован в {fn}")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка экспорта", str(e))

    def _export_project_to_excel(self, fn, result):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        wb = Workbook(); ws1 = wb.active; ws1.title = "Сводка проекта"
        hf = PatternFill(start_color="6C63FF", end_color="6C63FF", fill_type="solid")
        thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                      top=Side(style='thin'), bottom=Side(style='thin'))
        ws1.merge_cells('A1:F1')
        ws1['A1'].value = f"Проект: {result['city']} — {result['specialization']}"
        ws1['A1'].font = Font(bold=True, size=14)
        ws1['A1'].alignment = Alignment(horizontal='center')
        calc = result['calculations']; ip = result['input_params']
        for i, row in enumerate([
            ["Город:", result['city']], ["Специализация:", result['specialization']],
            ["Транспорт:", result['transport_type']],
            ["Всего визитов:", f"{ip['total_visits_needed']:,}"],
            ["Визитов на врача:", ip['visits_per_doctor']],
            ["Срок:", f"{ip['project_calendar_days']} дней"], [],
            ["Уникальных врачей:", f"{calc['unique_doctors_needed']:,}"],
            ["Мин. медпредов:", calc['min_reps_needed']],
            ["Оптим. медпредов:", calc.get('optimal_reps_needed', calc['min_reps_needed'])],
            ["Напряжённость:", f"{calc.get('project_status','—')} ({calc.get('project_intensity',0)}%)"]
        ], 3):
            if row:
                ws1[f'A{i}'] = row[0]; ws1[f'A{i}'].font = Font(bold=True)
                ws1[f'B{i}'] = row[1]
                ws1[f'A{i}'].border = thin; ws1[f'B{i}'].border = thin
        ws1.column_dimensions['A'].width = 30; ws1.column_dimensions['B'].width = 30
        if 'scenarios' in result and result['scenarios']:
            ws2 = wb.create_sheet("Варианты реализации")
            hdrs = ['Тип','Медпредов','Недель','Раб. дней','Календ. дней',
                    'Исп. сроков%','Загрузка%','Рекомендация']
            for ci, h in enumerate(hdrs, 1):
                cell = ws2.cell(row=1, column=ci, value=h)
                cell.font = Font(bold=True, color="FFFFFF"); cell.fill = hf
                cell.alignment = Alignment(horizontal='center')
            for ri, sc in enumerate(result['scenarios'], 2):
                tp = ("ОПТИМАЛЬНО" if sc.get('is_optimal') else
                      "МИНИМАЛЬНО" if sc.get('is_minimal') else "—")
                for ci, v in enumerate([
                    tp, sc['reps_count'], sc['weeks'], int(sc['work_days']),
                    int(sc['calendar_days']), sc['time_utilization'],
                    sc['rep_utilization'], sc.get('recommendation','')], 1):
                    ws2.cell(row=ri, column=ci, value=v)
            for col in ws2.columns:
                ml = max((len(str(c.value)) for c in col if c.value), default=0)
                ws2.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 3, 40)
        wb.save(fn)

    # ── Прочее ───────────────────────────────────────────────────────────────

    def reset_calculation(self):
        self.results_panel.metrics_widget.setVisible(False)
        self.results_panel.schedule_table.setRowCount(0)
        self.results_panel.recommendations_text.clear()
        self.results_panel.map_view.setHtml("")
        while self.results_panel.graph_layout.count():
            item = self.results_panel.graph_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        ph = QLabel("Выполните расчёт для отображения графиков")
        ph.setAlignment(Qt.AlignCenter)
        ph.setStyleSheet(f"color: {C['text3']}; font-size: 14px; padding: 60px;")
        self.results_panel.graph_layout.addWidget(ph)
        self.daily_calc_panel.export_btn.setEnabled(False)
        self.results_title.setText("Результаты расчёта")
        self.status_bar.showMessage("Готов к новому расчёту")
        self.status_dot.setText("●  Готов")
        self.status_dot.setStyleSheet(f"color: {C['success']}; font-size: 12px; font-weight: 600;")
        self.results_panel.set_current_tab(0)

    def show_city_statistics(self):
        try:
            df = self.calculator.get_city_statistics()
            dialog = QDialog(self)
            dialog.setWindowTitle("Статистика по городам")
            dialog.setMinimumSize(860, 500)
            dialog.setStyleSheet(f"background-color: {C['surface']}; color: {C['text']};")
            lay = QVBoxLayout(dialog); lay.setContentsMargins(16, 16, 16, 16)
            table = AppTable(len(df.columns), df.columns.tolist())
            table.setRowCount(len(df))
            for i, row in df.iterrows():
                for j, col in enumerate(df.columns):
                    table.set_cell(i, j, str(row[col]))
            table.resizeColumnsToContents(); lay.addWidget(table)
            bb = QDialogButtonBox(QDialogButtonBox.Ok)
            bb.accepted.connect(dialog.accept); lay.addWidget(bb)
            dialog.exec()
        except Exception as e:
            QMessageBox.warning(self, "Ошибка", f"Не удалось загрузить статистику:\n{str(e)}")

    def show_about(self):
        QMessageBox.about(self, "О программе",
            f"<h2>ELOMER</h2>"
            f"<p style='color:#888;'>Efficient LOgistics of a MEdical Representative</p>"
            f"<p>Версия 1.0 | Оптимизация работы медицинских представителей</p>"
            f"<p><b>Возможности:</b></p><ul>"
            f"<li>Расчёт рабочего дня и перемещений</li>"
            f"<li>Проектное планирование со сценариями</li>"
            f"<li>Анализ рисков методом Монте-Карло</li>"
            f"<li>Интерактивная карта маршрута</li>"
            f"<li>ML-оптимизация количества визитов</li>"
            f"<li>Экспорт в Excel</li></ul>"
            f"<p><i>Python 3 · PySide6 · Matplotlib · Scikit-learn · Folium</i></p>")

    def _create_simple_schedule(self, result):
        import datetime as dt
        schedule = []
        cur = dt.datetime.strptime('09:00', '%H:%M')
        vt = result.get('visit_times',  [30] * result['num_visits'])
        tt = result.get('travel_times', [20] * max(0, result['num_visits'] - 1))
        for i in range(result['num_visits']):
            vd = vt[i] if i < len(vt) else 30
            schedule.append({'time': cur.strftime('%H:%M'),
                              'activity': f"Визит {i+1}: {result['specialization']}",
                              'duration_min': round(vd, 1), 'type': 'visit'})
            cur += dt.timedelta(minutes=vd)
            if i < result['num_visits'] - 1:
                td = tt[i] if i < len(tt) else 20
                schedule.append({'time': cur.strftime('%H:%M'),
                                  'activity': f"Перемещение к точке {i+2}",
                                  'duration_min': round(td, 1), 'type': 'travel'})
                cur += dt.timedelta(minutes=td)
        admin = result.get('admin_time_min', 60)
        if admin > 0:
            schedule.append({'time': cur.strftime('%H:%M'),
                              'activity': 'Административная работа (отчёты)',
                              'duration_min': round(admin, 1), 'type': 'admin'})
        return schedule

    def _generate_simple_locations(self, city, spec, num_visits):
        coords = {
            'Москва': (55.7558, 37.6173), 'Санкт-Петербург': (59.9343, 30.3351),
            'Екатеринбург': (56.8389, 60.6057), 'Новосибирск': (55.0084, 82.9357),
            'Казань': (55.7961, 49.1064)
        }
        blat, blon = coords.get(city, (55.7558, 37.6173))
        return [{'id': i + 1,
                 'type': 'Аптека' if spec == 'Аптеки' else 'Поликлиника',
                 'name': f'{spec} {i+1}',
                 'latitude':  blat + random.uniform(-0.1,  0.1),
                 'longitude': blon + random.uniform(-0.15, 0.15),
                 'specialization': spec}
                for i in range(num_visits)]

    def closeEvent(self, event):
        reply = QMessageBox.question(self, "Выход", "Вы уверены, что хотите выйти?",
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        event.accept() if reply == QMessageBox.Yes else event.ignore()

    # ── Совместимость ────────────────────────────────────────────────────────
    def format_city_load_report(self, result):   return self._format_project_report(result)
    def update_scenarios_table(self, result, t): pass
    def update_project_scenarios_table(self, r): self._fill_project_scenarios_table(r)
    def show_project_results_in_tab(self, r):    self._show_project_results(r)
    def format_project_report(self, result):     return self._format_project_report(result)
    def calculate_city_load(self):               self.calculate_project()



# Алиасы обратной совместимости
ModernButton   = AppButton
ModernComboBox = AppComboBox
ModernSpinBox  = AppSpinBox
ModernCheckBox = AppCheckBox
InputPanel     = DailyCalcPanel
SectionLabel   = SectionDivider
FieldLabel     = QLabel
